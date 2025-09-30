#!/usr/bin/env python
"""
Excel 解析功能演示脚本

快速演示如何使用 Excel 解析功能
"""

import asyncio
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook


async def demo():
    """演示 Excel 解析功能"""

    print("📊 Excel 解析功能演示")
    print("="*60)

    # 1. 创建简单的测试 Excel 文件
    print("\n1️⃣  创建测试 Excel 文件...")

    wb = Workbook()
    ws = wb.active
    ws.title = "示例数据"

    # 添加数据
    ws['A1'] = "姓名"
    ws['B1'] = "年龄"
    ws['C1'] = "城市"

    data = [
        ("张三", 25, "北京"),
        ("李四", 30, "上海"),
        ("王五", 28, "深圳"),
    ]

    for i, (name, age, city) in enumerate(data, start=2):
        ws[f'A{i}'] = name
        ws[f'B{i}'] = age
        ws[f'C{i}'] = city

    # 添加公式
    ws['A5'] = "平均年龄"
    ws['B5'] = "=AVERAGE(B2:B4)"

    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)
    demo_file = output_dir / "demo_excel.xlsx"

    wb.save(str(demo_file))
    print(f"   ✅ 已创建: {demo_file}")

    # 2. 演示解析功能
    print("\n2️⃣  解析 Excel 文件...")
    print(f"\n   📄 文件: {demo_file}")
    print(f"   📊 工作表: {wb.sheetnames}")
    print(f"   📐 数据行: {ws.max_row}")
    print(f"   📏 数据列: {ws.max_column}")

    # 3. 使用说明
    print("\n3️⃣  使用方法:")
    print("""
   在 Claude Code 或其他 MCP 客户端中使用:

   >>> 请帮我解析这个 Excel 文件: output/demo_excel.xlsx

   或指定工作表:

   >>> 解析 output/demo_excel.xlsx 中的 "示例数据" 工作表

   或快速提取文本:

   >>> 提取 output/demo_excel.xlsx 的文本内容

   或获取元数据:

   >>> 获取 output/demo_excel.xlsx 的元数据信息
    """)

    print("\n✨ 演示完成!")
    print("="*60)


if __name__ == "__main__":
    asyncio.run(demo())