#!/usr/bin/env python
"""
测试 Excel 文档解析功能

此脚本创建测试 Excel 文件并测试解析功能
"""

import asyncio
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# 添加 src 目录到路径
sys.path.insert(0, str(Path(__file__).parent))

from src.server import DocxTemplateServer


def create_test_excel():
    """创建测试 Excel 文件"""
    wb = Workbook()

    # 第一个工作表：销售数据
    ws1 = wb.active
    ws1.title = "销售数据"

    # 添加标题
    ws1['A1'] = "产品名称"
    ws1['B1'] = "销售数量"
    ws1['C1'] = "单价"
    ws1['D1'] = "总计"

    # 设置标题样式
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 添加数据
    products = [
        ("笔记本电脑", 10, 5999, "=B2*C2"),
        ("鼠标", 50, 89, "=B3*C3"),
        ("键盘", 30, 299, "=B4*C4"),
        ("显示器", 15, 1899, "=B5*C5"),
    ]

    for i, (name, qty, price, formula) in enumerate(products, start=2):
        ws1[f'A{i}'] = name
        ws1[f'B{i}'] = qty
        ws1[f'C{i}'] = price
        ws1[f'D{i}'] = formula

    # 添加总计行
    ws1['A6'] = "总计"
    ws1['A6'].font = Font(bold=True)
    ws1['D6'] = "=SUM(D2:D5)"

    # 第二个工作表：员工信息
    ws2 = wb.create_sheet("员工信息")
    ws2['A1'] = "姓名"
    ws2['B1'] = "部门"
    ws2['C1'] = "入职日期"

    # 设置标题样式
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 添加员工数据
    employees = [
        ("张三", "技术部", datetime(2023, 1, 15)),
        ("李四", "市场部", datetime(2023, 3, 20)),
        ("王五", "财务部", datetime(2023, 6, 10)),
    ]

    for i, (name, dept, date) in enumerate(employees, start=2):
        ws2[f'A{i}'] = name
        ws2[f'B{i}'] = dept
        ws2[f'C{i}'] = date

    # 第三个工作表：统计数据
    ws3 = wb.create_sheet("统计数据")
    ws3['A1'] = "指标"
    ws3['B1'] = "数值"

    ws3['A2'] = "总销售额"
    ws3['B2'] = "=销售数据!D6"
    ws3['A3'] = "平均单价"
    ws3['B3'] = "=AVERAGE(销售数据!C2:C5)"
    ws3['A4'] = "总销售量"
    ws3['B4'] = "=SUM(销售数据!B2:B5)"

    # 合并单元格示例
    ws3.merge_cells('A5:B5')
    ws3['A5'] = "这是合并的单元格"
    ws3['A5'].alignment = Alignment(horizontal='center')

    # 保存文件
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)

    test_file = output_dir / "test_excel.xlsx"
    wb.save(str(test_file))

    print(f"✅ 测试 Excel 文件已创建: {test_file}")
    print(f"   - 工作表数: {len(wb.sheetnames)}")
    print(f"   - 工作表名称: {', '.join(wb.sheetnames)}")

    return test_file


async def test_parse_excel(server, file_path):
    """测试 Excel 解析功能"""

    print("\n" + "="*60)
    print("测试 1: 解析所有工作表")
    print("="*60)

    result = await server.parse_excel_document(
        file_path=str(file_path),
        include_formulas=True
    )

    print(result[0].text)

    print("\n" + "="*60)
    print("测试 2: 解析指定工作表")
    print("="*60)

    result = await server.parse_excel_document(
        file_path=str(file_path),
        sheet_name="销售数据",
        include_formulas=True
    )

    print(result[0].text)

    print("\n" + "="*60)
    print("测试 3: 提取文本 (快速模式)")
    print("="*60)

    result = await server.extract_text_from_document(
        file_path=str(file_path)
    )

    print(result[0].text)

    print("\n" + "="*60)
    print("测试 4: 提取元数据")
    print("="*60)

    result = await server.get_document_metadata(
        file_path=str(file_path)
    )

    print(result[0].text)


async def main():
    """主测试函数"""
    print("🧪 Excel 文档解析功能测试")
    print("="*60)

    # 创建测试文件
    test_file = create_test_excel()

    # 初始化服务器
    server = DocxTemplateServer()

    # 运行测试
    await test_parse_excel(server, test_file)

    print("\n" + "="*60)
    print("✅ 所有测试完成!")
    print("="*60)


if __name__ == "__main__":
    asyncio.run(main())